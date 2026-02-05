"""
FastAPI Backend for LME Data Fetcher
Uses Playwright to bypass Cloudflare protection
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from datetime import datetime
import io
import json
import os
from pathlib import Path
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import time

# Import database functions
from database import (
    save_commodity_prices,
    log_fetch_operation,
    get_all_commodity_data,
    get_latest_prices,
    get_data_summary,
    log_change_event
)
from pydantic import BaseModel
from typing import List

app = FastAPI(title="LME Data API")

# Data storage directory
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

# CORS configuration for Next.js frontend
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        FRONTEND_URL,
        "http://localhost:3000",
        "https://cv.ndtduy.live"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data Source Configuration
DATA_SOURCES = {
    "copper": {
        "type": "lme",
        "id": "39fabad0-95ca-491b-a733-bcef31818b16",
        "name": "Copper",
        "url": "https://www.lme.com/metals/non-ferrous/lme-copper"
    },
    "zinc": {
        "type": "lme",
        "id": "1a1aca59-3032-4ea6-b22b-18b151514b84",
        "name": "Zinc",
        "url": "https://www.lme.com/metals/non-ferrous/lme-zinc"
    },
    "oil": {
        "type": "oilprice",
        "name": "Oil Price (WTI)",
        "url": "https://oilprice.com/",
        "api_url": "https://oilprice.com/freewidgets/json_get_oilprices",
        "blend_id": 39,
        "period": 7
    }
}

START_DATE = "2023-01-01"

def get_end_date():
    """Get current date in YYYY-MM-DD format"""
    return datetime.now().strftime("%Y-%m-%d")

def convert_date_format(date_str: str) -> str:
    """Convert date from DD/MM/YYYY to YYYY-MM-DD for PostgreSQL"""
    try:
        if '/' in date_str:
            # DD/MM/YYYY format
            parts = date_str.split('/')
            if len(parts) == 3:
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return date_str
    except:
        return date_str

def save_data_to_file(source: str, data: dict):
    """Save fetched data to JSON file"""
    try:
        filename = DATA_DIR / f"{source}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        latest_filename = DATA_DIR / f"{source}_latest.json"

        with open(filename, 'w') as f:
            json.dump(data, f, indent=2)

        # Also save as latest
        with open(latest_filename, 'w') as f:
            json.dump(data, f, indent=2)

        print(f"[Storage] Saved data to {filename}")
    except Exception as e:
        print(f"[Storage] Error saving data: {str(e)}")

def load_data_from_file(source: str) -> dict:
    """Load latest data from file"""
    try:
        filename = DATA_DIR / f"{source}_latest.json"
        if filename.exists():
            with open(filename, 'r') as f:
                data = json.load(f)
            print(f"[Storage] Loaded data from {filename}")
            return data
        return None
    except Exception as e:
        print(f"[Storage] Error loading data: {str(e)}")
        return None

async def fetch_oil_price_data_with_playwright():
    """
    Fetch oil price data from oilprice.com using Playwright
    Fetches CSRF token first, then uses it to get price data
    """
    from playwright.async_api import async_playwright

    print("[Playwright] Fetching Oil Price data...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
        )
        page = await context.new_page()

        try:
            print("[Playwright] Navigating to oilprice.com...")
            # Use domcontentloaded instead of networkidle to avoid timeouts from ads/trackers
            await page.goto("https://oilprice.com/", wait_until="domcontentloaded", timeout=60000)

            print("[Playwright] Waiting for page to load...")
            await asyncio.sleep(2)


            print("[Playwright] Fetching CSRF token...")
            # Fetch CSRF token from the dedicated endpoint
            csrf_response = await page.evaluate("""
                async () => {
                    const response = await fetch('https://oilprice.com/ajax/csrf', {
                        method: 'GET',
                        headers: {
                            'accept': 'application/json, text/javascript, */*; q=0.01',
                            'x-requested-with': 'XMLHttpRequest'
                        }
                    });

                    if (!response.ok) {
                        throw new Error(`CSRF fetch failed: ${response.status}`);
                    }

                    const data = await response.json();
                    // Return the token directly as a string
                    return data.token || data.csrf_token || data;
                }
            """)

            # Extract CSRF token from response - it should already be a string
            csrf_token = str(csrf_response) if csrf_response else ""
            print(f"[Playwright] Got CSRF token: {csrf_token[:20] if len(csrf_token) > 20 else csrf_token}...")

            print(f"[Playwright] Making API request with CSRF token...")

            # Make POST request from browser context with the fetched CSRF token
            response_data = await page.evaluate("""
                async (csrf) => {
                    const response = await fetch('https://oilprice.com/freewidgets/json_get_oilprices', {
                        method: 'POST',
                        headers: {
                            'accept': 'application/json, text/javascript, */*; q=0.01',
                            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
                            'x-requested-with': 'XMLHttpRequest'
                        },
                        body: `blend_id=39&period=7&op_csrf_token=${csrf}&futures=1`
                    });

                    if (!response.ok) {
                        throw new Error(`HTTP ${response.status}: ${response.statusText}`);
                    }

                    return await response.json();
                }
            """, csrf_token)

            print(f"[Playwright] Successfully fetched oil price data!")

            await browser.close()

            return response_data

        except Exception as e:
            await browser.close()
            print(f"[Playwright] Error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Failed to fetch oil price data: {str(e)}")

async def fetch_lme_data_with_playwright(start_date: str, end_date: str, source: str = "copper"):
    """
    Use Playwright to fetch LME data directly from browser session
    This bypasses Cloudflare by making requests from a real browser
    """
    from playwright.async_api import async_playwright

    # Get source configuration
    if source not in DATA_SOURCES:
        raise HTTPException(status_code=400, detail=f"Invalid source: {source}. Available: {', '.join(DATA_SOURCES.keys())}")

    source_config = DATA_SOURCES[source]
    datasource_id = source_config["id"]
    source_url = source_config["url"]
    source_name = source_config["name"]

    print(f"[Playwright] Fetching {source_name} data...")
    print("[Playwright] Launching browser to bypass Cloudflare...")

    async with async_playwright() as p:
        # Launch browser in headless mode
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
        )
        page = await context.new_page()

        try:
            print(f"[Playwright] Navigating to {source_name} page...")
            await page.goto(source_url, wait_until="networkidle", timeout=60000)

            print("[Playwright] Waiting for page to load...")
            await asyncio.sleep(3)

            # Try to accept cookies consent if present
            try:
                accept_button = page.locator('button:has-text("Accept")')
                if await accept_button.count() > 0:
                    print("[Playwright] Accepting cookies consent...")
                    await accept_button.first.click()
                    await asyncio.sleep(1)
            except:
                pass

            # Scroll to trigger any lazy-loaded content
            print("[Playwright] Scrolling page...")
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(2)

            # Now make the API request from the browser context
            print(f"[Playwright] Fetching {source_name} data from {start_date} to {end_date}...")
            api_url = f"https://www.lme.com/api/trading-data/chart-data?datasourceId={datasource_id}&startDate={start_date}&endDate={end_date}"

            # Use page.evaluate to make the fetch call from the browser context
            # This bypasses Cloudflare because it's a real browser making the request
            response_data = await page.evaluate("""
                async (url) => {
                    const response = await fetch(url, {
                        method: 'GET',
                        headers: {
                            'accept': '*/*',
                            'cache-control': 'no-cache'
                        }
                    });

                    if (!response.ok) {
                        throw new Error(`HTTP ${response.status}: ${response.statusText}`);
                    }

                    return await response.json();
                }
            """, api_url)

            print(f"[Playwright] Successfully fetched data! {len(response_data.get('Labels', []))} records")

            await browser.close()

            return response_data

        except Exception as e:
            await browser.close()
            print(f"[Playwright] Error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Failed to fetch data: {str(e)}")

class ChangeLogRequest(BaseModel):
    summary: str
    details: List[str]

@app.post("/api/logs/change")
async def create_change_log(request: ChangeLogRequest):
    """Log a data change event"""
    try:
        success = log_change_event(request.summary, request.details)
        if not success:
            raise HTTPException(status_code=500, detail="Failed to save log to database")
        return {"status": "success", "message": "Log saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def create_oil_price_excel(data: dict) -> bytes:
    """Create Excel file from oil price data"""

    print(f"[Excel] Creating Oil Price Excel file")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Oil Prices"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Extract price history from data
    # Oil price API returns different structure
    prices = data.get("prices", [])

    if not prices:
        raise HTTPException(status_code=400, detail="No price data received from API")

    # Write headers
    headers = ["Date", "Price (USD)", "Change"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, price_data in enumerate(prices, start=2):
        ws.cell(row=row_idx, column=1, value=price_data.get("date", ""))
        ws.cell(row=row_idx, column=2, value=price_data.get("price", ""))
        ws.cell(row=row_idx, column=3, value=price_data.get("change", ""))

    # Auto-adjust column widths
    for col_idx in range(1, 4):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()

def create_excel_file(data: dict) -> bytes:
    """Create Excel file from LME data"""

    labels = data.get("Labels", [])
    datasets = data.get("Datasets", [])

    if not labels:
        raise HTTPException(status_code=400, detail="No data received from LME API")

    print(f"[Excel] Creating Excel with {len(labels)} records")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Official Prices"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Write headers
    headers = ["Date"]
    for dataset in datasets:
        row_title = dataset.get("RowTitle", "Unknown")
        label = dataset.get("Label", "")
        column_name = f"{row_title} - {label}" if label else row_title
        headers.append(column_name)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, label in enumerate(labels, start=2):
        ws.cell(row=row_idx, column=1, value=label)

        for col_idx, dataset in enumerate(datasets, start=2):
            data_values = dataset.get("Data", [])
            value = data_values[row_idx - 2] if row_idx - 2 < len(data_values) else None
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(2, min(len(labels) + 2, 100)):  # Check first 100 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()

@app.get("/api/get-cached-data")
async def get_cached_data(source: str = "copper"):
    """
    Get cached data from file (if available)

    Parameters:
    - source: copper, zinc, or oil (default: copper)
    """
    try:
        if source not in DATA_SOURCES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid source: {source}. Available: {', '.join(DATA_SOURCES.keys())}"
            )

        data = load_data_from_file(source)

        if data is None:
            raise HTTPException(
                status_code=404,
                detail=f"No cached data found for {source}. Please fetch fresh data first."
            )

        return data

    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"[API Cached] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/db/chart-data")
async def get_chart_data_from_db(start_date: str = "2023-01-01"):
    """
    Get all commodity data from database for chart display

    Parameters:
    - start_date: Start date in YYYY-MM-DD format (default: 2023-01-01)
    """
    try:
        print(f"[API DB] Loading chart data from {start_date}...")
        data = get_all_commodity_data(start_date)

        # Transform to chart format
        chart_data_map = {}

        for item in data:
            date = item['date']
            code = item['commodity_code'].lower()

            if date not in chart_data_map:
                chart_data_map[date] = {'date': date}

            chart_data_map[date][code] = item['price_value']

        # Convert to array and sort
        chart_data = sorted(chart_data_map.values(), key=lambda x: x['date'])

        return {
            "success": True,
            "data": chart_data,
            "count": len(chart_data),
            "loaded_at": datetime.now().isoformat()
        }

    except Exception as e:
        print(f"[API DB] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/db/latest-prices")
async def get_latest_from_db():
    """Get latest prices for all commodities from database"""
    try:
        data = get_latest_prices()

        return {
            "success": True,
            "data": data,
            "count": len(data),
            "loaded_at": datetime.now().isoformat()
        }

    except Exception as e:
        print(f"[API DB] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/db/summary")
async def get_db_summary():
    """Get database summary (record counts, date ranges)"""
    try:
        summary = get_data_summary()

        return {
            "success": True,
            "summary": summary,
            "loaded_at": datetime.now().isoformat()
        }

    except Exception as e:
        print(f"[API DB] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
def read_root():
    """Health check endpoint"""
    return {
        "status": "running",
        "service": "Commodity Data API",
        "version": "3.0.0",
        "method": "Direct Playwright fetch (Cloudflare bypass)",
        "database": "PostgreSQL - commodity_data",
        "endpoints": {
            "fetch": "/api/fetch-data-json?source={copper|zinc|oil}",
            "db_chart": "/api/db/chart-data?start_date=2023-01-01",
            "db_latest": "/api/db/latest-prices",
            "db_summary": "/api/db/summary",
            "cached_file": "/api/get-cached-data?source={copper|zinc|oil}",
            "download": "/api/fetch-lme-data-direct?source={copper|zinc|oil}"
        },
        "available_sources": list(DATA_SOURCES.keys()),
        "sources": {k: v["name"] for k, v in DATA_SOURCES.items()},
        "storage": "PostgreSQL + JSON files"
    }

@app.get("/api/fetch-data-json")
async def fetch_data_json(source: str = "copper"):
    """
    Fetch data as JSON (for chart display)

    Parameters:
    - source: copper, zinc, or oil (default: copper)
    """

    try:
        # Validate source
        if source not in DATA_SOURCES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid source: {source}. Available: {', '.join(DATA_SOURCES.keys())}"
            )

        source_config = DATA_SOURCES[source]
        source_name = source_config["name"]
        source_type = source_config["type"]

        # Track fetch time
        start_time = time.time()

        # Fetch data based on type
        if source_type == "lme":
            print(f"[API JSON] Fetching {source_name} data from {START_DATE} to {get_end_date()}")
            data = await fetch_lme_data_with_playwright(START_DATE, get_end_date(), source)

            # Transform LME data for chart - parse Labels and Datasets structure
            chart_data = []
            db_data = []

            labels = data.get("Labels", [])
            datasets = data.get("Datasets", [])

            # Build a map of dataset values by row title for easy lookup
            dataset_map = {}
            for dataset in datasets:
                row_title = dataset.get("RowTitle", "")
                label = dataset.get("Label", "")
                key = f"{row_title}_{label}".lower().replace(" ", "_")
                dataset_map[key] = dataset.get("Data", [])

            # Extract specific datasets (Cash Bid, Cash Offer, 3-Month Bid, 3-Month Offer)
            cash_bid_data = dataset_map.get("cash_bid", dataset_map.get("official_price_bid", []))
            cash_offer_data = dataset_map.get("cash_offer", dataset_map.get("official_price_offer", []))
            three_month_bid_data = dataset_map.get("3-months_bid", dataset_map.get("3_months_bid", []))
            three_month_offer_data = dataset_map.get("3-months_offer", dataset_map.get("3_months_offer", []))

            print(f"[API JSON] Processing {len(labels)} dates with datasets: {list(dataset_map.keys())}")

            for i, date_label in enumerate(labels):
                cash_bid = cash_bid_data[i] if i < len(cash_bid_data) else None
                cash_offer = cash_offer_data[i] if i < len(cash_offer_data) else None
                three_month_bid = three_month_bid_data[i] if i < len(three_month_bid_data) else None
                three_month_offer = three_month_offer_data[i] if i < len(three_month_offer_data) else None

                # Use cash bid as primary value, fallback to 3-month bid
                value = cash_bid or three_month_bid

                chart_item = {
                    "date": date_label,
                    "value": value,
                    "source": source_name
                }
                chart_data.append(chart_item)

                # Prepare for database (convert date format for PostgreSQL)
                db_date = convert_date_format(date_label)
                db_data.append({
                    "date": db_date,
                    "value": value,
                    "cashBid": cash_bid,
                    "cashOffer": cash_offer,
                    "threeMonthBid": three_month_bid,
                    "threeMonthOffer": three_month_offer,
                    "source": "LME"
                })

            # Calculate duration
            duration_ms = int((time.time() - start_time) * 1000)

            # Save to database
            try:
                records_saved = save_commodity_prices(source.upper(), db_data)
                log_fetch_operation(source.upper(), "success", records_saved, None, duration_ms)
                print(f"[DB] Saved {records_saved} records to database")
            except Exception as db_error:
                print(f"[DB] Error saving to database: {str(db_error)}")
                log_fetch_operation(source.upper(), "partial", 0, str(db_error), duration_ms)

            result = {
                "source": source,
                "name": source_name,
                "data": chart_data,
                "fetched_at": datetime.now().isoformat(),
                "saved_to_db": records_saved if 'records_saved' in locals() else 0
            }

            # Save to file
            save_data_to_file(source, result)

            return result

        elif source_type == "oilprice":
            print(f"[API JSON] Fetching {source_name} data")
            data = await fetch_oil_price_data_with_playwright()

            # Transform oil price data for chart
            chart_data = []
            db_data = []
            prices = data.get("prices", [])
            for item in prices:
                # Handle timestamp from 'time' field (Unix timestamp in seconds)
                date_str = None
                if item.get("time"):
                    try:
                        date_str = datetime.fromtimestamp(int(item.get("time"))).strftime("%Y-%m-%d")
                    except:
                        pass

                # Skip dates before START_DATE to match LME data range
                if date_str and date_str < START_DATE:
                    continue

                chart_item = {
                    "date": date_str,
                    "value": item.get("price"),
                    "source": source_name
                }
                chart_data.append(chart_item)

                # Prepare for database
                if date_str:
                    db_data.append({
                        "date": date_str,
                        "value": item.get("price"),
                        "source": "OilPrice.com"
                    })

            # Calculate duration
            duration_ms = int((time.time() - start_time) * 1000)

            # De-duplicate db_data by date (keep latest value for each date)
            unique_db_data = {}
            for item in db_data:
                unique_db_data[item["date"]] = item
            db_data = list(unique_db_data.values())
            print(f"[Oil] De-duplicated to {len(db_data)} unique records")

            # Save to database
            try:
                records_saved = save_commodity_prices("OIL", db_data)
                log_fetch_operation("OIL", "success", records_saved, None, duration_ms)
                print(f"[DB] Saved {records_saved} records to database")
            except Exception as db_error:
                print(f"[DB] Error saving to database: {str(db_error)}")
                log_fetch_operation("OIL", "partial", 0, str(db_error), duration_ms)

            result = {
                "source": source,
                "name": source_name,
                "data": chart_data,
                "fetched_at": datetime.now().isoformat(),
                "saved_to_db": records_saved if 'records_saved' in locals() else 0
            }

            # Save to file
            save_data_to_file(source, result)

            return result

        else:
            raise HTTPException(status_code=400, detail=f"Unknown source type: {source_type}")

    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"[API JSON] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/fetch-lme-data-direct")
async def fetch_data_direct(source: str = "copper"):
    """
    Fetch data directly using Playwright (100% Cloudflare bypass)
    No cookies needed - uses real browser session

    Parameters:
    - source: copper, zinc, or oil (default: copper)
    """

    try:
        # Validate source
        if source not in DATA_SOURCES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid source: {source}. Available: {', '.join(DATA_SOURCES.keys())}"
            )

        source_config = DATA_SOURCES[source]
        source_name = source_config["name"]
        source_type = source_config["type"]

        # Fetch data based on type
        if source_type == "lme":
            print(f"[API] Starting {source_name} direct fetch from {START_DATE} to {get_end_date()}")
            data = await fetch_lme_data_with_playwright(START_DATE, get_end_date(), source)
            excel_bytes = create_excel_file(data)
            filename = f"LME_{source_name}_Official_Prices_{datetime.now().strftime('%Y_%m_%d')}.xlsx"

        elif source_type == "oilprice":
            print(f"[API] Starting {source_name} direct fetch")
            data = await fetch_oil_price_data_with_playwright()
            excel_bytes = create_oil_price_excel(data)
            filename = f"Oil_Price_WTI_{datetime.now().strftime('%Y_%m_%d')}.xlsx"

        else:
            raise HTTPException(status_code=400, detail=f"Unknown source type: {source_type}")

        # Return as downloadable file
        print(f"[API] Returning Excel file: {filename}")

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"[API] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    print("=" * 60)
    print("LME Data API - Playwright Direct Fetch")
    print("=" * 60)
    print(f"Starting server at: http://localhost:8000")
    print(f"Data range: {START_DATE} to {get_end_date()}")
    print(f"Frontend: http://localhost:3000")
    print(f"Method: Direct browser fetch (100% Cloudflare bypass)")
    print("=" * 60)
    print()

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
